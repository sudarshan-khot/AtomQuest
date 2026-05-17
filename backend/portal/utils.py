"""
Utility functions for the AtomQuest Goal Setting & Tracking Portal.
"""
import json
from datetime import datetime, timedelta
from django.utils import timezone
from .models import AuditLog, Notification


# ============================================================================
# Audit Trail Utilities
# ============================================================================

def get_client_ip(request):
    """Get client IP address from request."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def log_audit_trail(entity_type, entity_id, action, user=None, old_values=None, 
                    new_values=None, comments='', ip_address='', user_agent=''):
    """Log an action to the audit trail."""
    AuditLog.objects.create(
        entity_type=entity_type,
        entity_id=entity_id,
        action=action,
        user=user,
        old_values=old_values,
        new_values=new_values,
        comments=comments,
        ip_address=ip_address,
        user_agent=user_agent
    )


# ============================================================================
# Notification Utilities
# ============================================================================

def send_notification(user, title, message, notification_type, goal=None, checkin=None):
    """Send a notification to a user."""
    Notification.objects.create(
        user=user,
        title=title,
        message=message,
        notification_type=notification_type,
        goal=goal,
        checkin=checkin
    )


# ============================================================================
# Progress Scoring Utilities
# ============================================================================

def calculate_progress_percentage(goal, progress_value):
    """
    Calculate progress percentage based on UoM type.
    
    Supports:
    - Numeric: (Current / Target) × 100, capped at 100%
    - Percentage: Direct mapping (0-100%)
    - Timeline: ((Current Date - Start) / (End - Start)) × 100
    - Zero-based: 0% if zero, 100% if non-zero
    """
    if not goal.uom_type:
        return 0
    
    uom_name = goal.uom_type.name
    
    if uom_name == 'numeric':
        if goal.target_value > 0:
            percentage = (progress_value / goal.target_value) * 100
            return min(percentage, 100)
        return 0
    
    elif uom_name == 'percentage':
        return min(max(progress_value, 0), 100)
    
    elif uom_name == 'timeline':
        # For timeline, progress_value should be current date
        # This is handled differently - typically in check-in creation
        return 0
    
    elif uom_name == 'zero_based':
        return 100 if progress_value > 0 else 0
    
    return 0


def calculate_timeline_progress(start_date, end_date, current_date=None):
    """
    Calculate progress percentage for timeline UoM.
    
    Formula: ((Current Date - Start) / (End - Start)) × 100
    """
    if current_date is None:
        current_date = timezone.now().date()
    
    # Convert to dates if needed
    if hasattr(start_date, 'date'):
        start_date = start_date.date()
    if hasattr(end_date, 'date'):
        end_date = end_date.date()
    if hasattr(current_date, 'date'):
        current_date = current_date.date()
    
    # Handle edge cases
    if current_date <= start_date:
        return 0
    if current_date >= end_date:
        return 100
    
    total_days = (end_date - start_date).days
    elapsed_days = (current_date - start_date).days
    
    if total_days == 0:
        return 0
    
    percentage = (elapsed_days / total_days) * 100
    return min(percentage, 100)


# ============================================================================
# Validation Utilities
# ============================================================================

def validate_weightage(goals):
    """
    Validate that total weightage equals 100%.
    
    Args:
        goals: List of Goal objects or queryset
    
    Returns:
        Tuple: (is_valid, total_weightage)
    """
    total = sum(goal.weightage for goal in goals)
    return total == 100, total


def validate_goal_count(user, cycle, exclude_goal_id=None):
    """
    Validate that user doesn't exceed max 8 goals per cycle.
    
    Args:
        user: User object
        cycle: Cycle object
        exclude_goal_id: Goal ID to exclude from count (for updates)
    
    Returns:
        Tuple: (is_valid, current_count, max_count)
    """
    from .models import Goal
    
    query = Goal.objects.filter(user=user, cycle=cycle)
    if exclude_goal_id:
        query = query.exclude(id=exclude_goal_id)
    
    count = query.count()
    max_count = 8
    
    return count < max_count, count, max_count


def validate_checkin_period(cycle, current_date=None):
    """
    Validate if current date is within an active check-in period.
    
    Args:
        cycle: Cycle object
        current_date: Date to check (defaults to today)
    
    Returns:
        Tuple: (is_active, next_checkin_date, period_name)
    """
    if current_date is None:
        current_date = timezone.now().date()
    
    # Define check-in periods (7 days before and after check-in date)
    periods = [
        ('Q1', cycle.checkin_date_q1),
        ('Q2', cycle.checkin_date_q2),
        ('Q3', cycle.checkin_date_q3),
        ('Q4', cycle.checkin_date_q4),
    ]
    
    for period_name, checkin_date in periods:
        if checkin_date:
            period_start = checkin_date - timedelta(days=7)
            period_end = checkin_date + timedelta(days=7)
            
            if period_start <= current_date <= period_end:
                return True, checkin_date, period_name
    
    # Find next check-in date
    next_checkin = None
    for period_name, checkin_date in periods:
        if checkin_date and checkin_date > current_date:
            if next_checkin is None or checkin_date < next_checkin:
                next_checkin = checkin_date
    
    return False, next_checkin, None


# ============================================================================
# Report Utilities
# ============================================================================

def calculate_weighted_achievement(goals_with_progress):
    """
    Calculate weighted achievement score across all goals.
    
    Args:
        goals_with_progress: List of tuples (goal, progress_percentage)
    
    Returns:
        Weighted achievement percentage
    """
    if not goals_with_progress:
        return 0
    
    total_weightage = sum(goal.weightage for goal, _ in goals_with_progress)
    if total_weightage == 0:
        return 0
    
    weighted_sum = sum(
        (progress / 100) * goal.weightage 
        for goal, progress in goals_with_progress
    )
    
    return (weighted_sum / total_weightage) * 100


# ============================================================================
# Data Export Utilities
# ============================================================================

def export_to_csv(data, filename=None):
    """
    Export data to CSV format.
    
    Args:
        data: List of dictionaries
        filename: Optional filename (without extension)
    
    Returns:
        CSV content as string
    """
    import csv
    from io import StringIO
    
    if not data:
        return ""
    
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)
    
    return output.getvalue()


def export_to_excel(data, filename=None):
    """
    Export data to Excel format.
    
    Args:
        data: List of dictionaries
        filename: Optional filename (without extension)
    
    Returns:
        Excel file content as bytes
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        if not data:
            return None
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Write headers
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Write data
        for row_num, row_data in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=row_data.get(header))
        
        # Auto-adjust column widths
        for col_num, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row_data in data:
                max_length = max(max_length, len(str(row_data.get(header, ''))))
            ws.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 50)
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
